#!/usr/bin/env python3

import sys
import pandas as pd
import openpyxl as op

PATH = sys.argv[1]

# Activating Excel Sheet Access
wb = op.load_workbook(PATH)
ws = wb.active

# Dyanmically setting data boundaries within the excel sheet
def left_bound_find():
    i = 1
    while True:
        value = ws.cell(row=1,column=i).value
        try:
            pd.to_datetime(value)
            return i
        except:
            i+=1

def right_bound_find():
    i = left_bound_find() + 1
    while True:
        value = ws.cell(row=1,column=i).value
        if pd.to_datetime(value) is None:
            return i
        i += 1

def lower_bound_find():
    i = 2
    while True:
        value = ws.cell(row=i, column=1).value
        if value is None:
            return i
        i += 1

upper_bound = 2  #Data will always begin on the second row
lower_bound = lower_bound_find()

left_bound = left_bound_find()
right_bound = right_bound_find()

id_col_bound = left_bound - 1 # translating excel column number to python indexing

# Forecast Setup and Melting
df = pd.read_excel(PATH)
ids = df.columns[:id_col_bound]
dates = df.columns[id_col_bound:]
df = df.fillna(0)
df_melt = pd.melt(
    df,
    id_vars=ids,
    value_vars=dates,
    var_name = 'Delivery Date',
    value_name = 'Order Quantity')
# Creating new melted table without order quanities of 0
index_drop = df_melt[df_melt['Order Quantity']==0].index
df_melt_short = df_melt.drop(index_drop).reset_index(drop=True)
# instantiating status column
df_melt_short['Status'] = ''

# Updating df_melt_short Status Column based on the excel sheet cell attributes
for i in range(df_melt_short.shape[0]):
    
    row_name = df_melt_short['New SKU'][i]
    col_name = df_melt_short['Delivery Date'][i]
    
    excel_rows = range(upper_bound,lower_bound)
    excel_columns = range(left_bound,right_bound)
    
    for row in excel_rows:
        for column in excel_columns:
            row_name_match = str(ws.cell(row=row, column = 1).value) == str(row_name)
            col_name_match = pd.to_datetime(ws.cell(row=1, column = column).value) == col_name
            if row_name_match and col_name_match:
                
                rgb = ws.cell(row=row, column = column).fill.fgColor.rgb
                tint = ws.cell(row=row, column = column).fill.fgColor.tint
                
                if rgb == 'FFFFFF00':
                    df_melt_short.at[i,'Status'] = 'Open'
                
                elif tint > 0:
                    df_melt_short.at[i,'Status'] = "+1.7m"
                    
                elif rgb == '00000000':
                    df_melt_short.at[i,'Status'] = "Recommended"

# export status-updated melt table to an .xlsx file
new_file_name = '.'.join(PATH.split('.')[:-1]) + ' - MELTED.xlsx'
df_melt_short.to_excel(new_file_name)